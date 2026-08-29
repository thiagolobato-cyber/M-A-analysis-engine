#!/usr/bin/env python3
"""
Roda um agente (Extraction, os 4 analíticos, ou CFO Synthesis) contra um deal.

Uso:
    python scripts/run_agent.py --agent complexity --deal-id <uuid>

O que faz:
  1. Busca no Supabase o deal_data (o "Data Object" canônico) do deal.
  2. Busca a versão ativa do agente e o system_prompt correspondente.
  3. Monta o prompt final e chama `claude -p` (autenticado via
     CLAUDE_CODE_OAUTH_TOKEN, já configurado como GitHub Secret).
  4. Valida que a resposta é um JSON bem formado.
  5. Grava o resultado em agent_runs (ou synthesis_runs, se for o CFO).

Nota: este é um primeiro rascunho. Testei a sintaxe e a lógica localmente,
mas não contra o Supabase real (esta sandbox não tem acesso de rede até
supabase.co) — vamos validar isso de verdade assim que os secrets do
GitHub Actions existirem.
"""
import argparse
import base64
import hashlib
import io
import json
import os
import re
import statistics
import subprocess
import sys
import urllib.request
import urllib.error
from datetime import datetime, timezone

from openpyxl import load_workbook

from dre_balancete_parser import (
    detect_consolidated_balancete,
    parse_consolidated_balancete,
    detect_dre_sheet,
    parse_dre_sheet,
    calcular_ebitda_de_dre,
    dre_linhas_para_contas,
    extrair_hierarquia_dre,
    calcular_resultado_de_hierarquia,
    extrair_margem_bruta_de_dre,
    montar_mini_dre,
    montar_tabela_viabilidade_financeira,
    agrupar_dre_linhas_por_trimestre,
    agregar_linhas_por_trimestre,
    _rotulos_legiveis_periodo,
)
from financial_engine import (
    mapear_waterfall,
    calcular_ebitda_3_camadas,
    detectar_anomalias_run_rate,
    detectar_contas_novas_ou_zeradas,
)
from complexity_rules import classificar_complexidade
from formulario_mapper import detectar_e_extrair_formulario, mapear_formulario
from regras_negocio import avaliar_viabilidade_financeira, avaliar_complexidade_operacional, avaliar_riscos_operacionais, avaliar_riscos_integracao, calcular_margem_bruta, gerar_red_flag_erp

MAX_ROWS_PER_SHEET = 300
MAX_COLS_PER_SHEET = 40
MAX_SHEETS = 30
MONTHLY_SHEET_RE = re.compile(r'(?i)^m[eê]s\s*0?(\d{1,2})$')

# Só o agente que lê DRE/Balancete de verdade precisa da série contábil
# crua — Complexity, Opinion, Integration/Operational Risks decidem a
# partir do resumo estruturado (geografia, sistema, HC, Pareto, etc.),
# nunca abriram a tabela de centenas de contas x 12 meses mesmo antes.
AGENTS_NEED_RAW_SERIES = {"financial_analysis"}

# Quantas contas (das potencialmente 300+) realmente vão pro prompt do
# agente financeiro — as demais já foram descartadas por não serem nem
# materiais nem voláteis o suficiente pra merecer leitura humana ou de IA.
MAX_ACCOUNTS_FOR_FINANCIAL_AGENT = 25

# Alias de modelo por padrão quando agent_versions.model vier vazio —
# nunca deixar cair no default da sessão (pode ser Opus sem ninguém notar).
DEFAULT_MODEL_ALIAS = "sonnet"


def detect_monthly_sheets(wb):
    """Detecta abas no padrão 'MES 01'..'MES 12' — um mês por aba, mesmo
    layout de colunas em cada uma. Comum em exports de Balancete brasileiros."""
    found = []
    for name in wb.sheetnames:
        m = MONTHLY_SHEET_RE.match(name.strip())
        if m:
            found.append((int(m.group(1)), name))
    found.sort()
    return found if len(found) >= 3 else []


def merge_monthly_balancete(wb, monthly_sheets):
    """Cruza as abas mensais por conta (nome + código), pegando o último
    valor numérico de cada linha como saldo final do mês — validado contra
    um arquivo real: Saldo Anterior + Débito − Crédito = Saldo Atual,
    sempre na última posição."""
    accounts = {}
    for month_num, sheet_name in monthly_sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if len(row) < 5:
                continue
            _, nome, codigo, tipo = row[0], row[1], row[2], row[3]
            if nome is None or codigo is None:
                continue
            numeric_vals = [v for v in row[4:] if isinstance(v, (int, float))]
            if not numeric_vals:
                continue
            key = (str(nome).strip(), codigo)
            if key not in accounts:
                accounts[key] = {"tipo": tipo, "meses": {}}
            accounts[key]["meses"][month_num] = numeric_vals[-1]
    return accounts


def format_merged_table(accounts, max_rows=300):
    """Formata o cruzamento como tabela, ordenada por relevância (maior
    valor absoluto em qualquer mês primeiro) — isso substitui pedir pro
    modelo cruzar 12 abas manualmente, o que não é confiável."""
    def materiality(item):
        return max((abs(v) for v in item[1]["meses"].values()), default=0)
    sorted_accounts = sorted(accounts.items(), key=materiality, reverse=True)

    lines = [
        "===== SÉRIE MENSAL CONSOLIDADA (cruzada automaticamente das abas mensais) =====",
        "Saldo final de cada mês, ordenado por relevância (maior valor absoluto primeiro).",
        "ATENÇÃO: contas de resultado (receita/despesa, geralmente código iniciando em 3 ou 4) "
        "costumam vir ACUMULADAS no ano-calendário — para o valor do mês isolado, calcule a "
        "diferença entre um mês e o anterior. Contas de balanço (ativo/passivo) já são o saldo "
        "no fim daquele mês, sem precisar de ajuste.",
        "Conta | Código | Tipo(S=sintética/A=analítica) | " + " | ".join(f"Mês{m:02d}" for m in range(1, 13)),
    ]
    for i, ((nome, codigo), data) in enumerate(sorted_accounts):
        if i >= max_rows:
            lines.append(f"[... {len(sorted_accounts) - max_rows} contas menos relevantes omitidas ...]")
            break
        valores = " | ".join(f"{data['meses'].get(m, 0):.2f}" if m in data["meses"] else "—" for m in range(1, 13))
        lines.append(f"{nome} | {codigo} | {data['tipo']} | {valores}")
    return "\n".join(lines)


def format_dre_table(dre_linhas: dict, periodos_rotulos: list | None = None, granularidade: str = "mensal") -> str:
    """Formata a DRE já extraída em código como texto compacto pro
    contexto — como a tabela já é pequena (dezenas de linhas, não
    centenas), incluir o texto inteiro aqui é barato; o que evitamos é
    o modelo precisar RECONSTRUIR isso a partir da aba genérica.

    BUG REAL CORRIGIDO EM 26/08 (achado pelo Thiago revisando o PPT de
    um deal real): esta função sempre rotulava as colunas como "Mês01",
    "Mês02"... mesmo quando os PERÍODOS reais eram "2025 (Ano Completo)"
    e "2026 (Q1 Jan-Mar)" — o agente `financial_analysis`, vendo só
    "Mês01: R$11,4M" e "Mês02: R$5,5M" no prompt, concluiu (corretamente,
    dado o que via) que a receita "caiu 51% entre dois meses consecutivos"
    e reportou isso como red flag/pergunta de due diligence — um achado
    FALSO: não são dois meses, é um ano inteiro comparado a um trimestre,
    onde a queda "aparente" é só efeito de escala de período, não
    variação de negócio. Corrigido: usa os rótulos reais quando
    disponíveis, e adiciona um aviso explícito quando a granularidade
    não é mensal, pra nunca mais o agente comparar período como se fosse
    mês a mês sem que isso seja verdade."""
    if periodos_rotulos and any(periodos_rotulos):
        col_labels = [r or f"mes_{m:02d}" for m, r in enumerate(periodos_rotulos, start=1)]
    else:
        col_labels = [f"Mês{m:02d}" for m in range(1, 13)]
    lines = [
        "===== DRE (extraída automaticamente pelo código, por rótulo de linha) =====",
        "Já vem pronta — não precisa copiar de volta no seu output, use como contexto.",
    ]
    if granularidade != "mensal":
        lines.append(
            f"ATENÇÃO — granularidade '{granularidade}', NÃO mensal: as colunas abaixo são "
            "períodos de duração DIFERENTE entre si (ex.: um ano completo vs um trimestre) — "
            "NUNCA trate como meses consecutivos nem calcule 'variação de um mês pro outro'. "
            "Se comparar períodos, normalize primeiro (ex.: anualizar o trimestre) ou compare "
            "só em %, nunca em R$ absoluto direto."
        )
    lines.append("Rótulo | " + " | ".join(col_labels))
    for rotulo, valores in dre_linhas.items():
        valores_fmt = " | ".join(
            f"{valores.get(f'mes_{m:02d}', 0):.2f}" if f"mes_{m:02d}" in valores else "—"
            for m in range(1, len(col_labels) + 1)
        )
        lines.append(f"{rotulo} | {valores_fmt}")
    return "\n".join(lines)


_PROMPT_CLASSIFICACAO_RAIZES = """Você recebe uma lista curta de rótulos de linha de uma DRE (Demonstração
de Resultado) — cada um é o nome de uma categoria de conta (ex.: "Marketing",
"IMPOSTOS", "Administrativas"), sem contexto de valor numérico.

Classifique CADA rótulo em exatamente um destes 3 tipos:
- "receita": entrada de dinheiro (venda, faturamento, prestação de serviço).
- "despesa": saída de dinheiro (custo, gasto, imposto, tributo, investimento).
- "resultado": linha de resultado calculado (subtotal, margem, lucro, EBITDA)
  — não é receita nem despesa em si, é derivada de outras linhas.

Se o rótulo for genuinamente ambíguo mesmo com bom senso de negócio
brasileiro (nunca invente contexto que não está no texto), classifique como
"despesa" por padrão — é a categoria mais comum numa DRE e o custo de errar
pro lado de despesa é menor que inflar receita por engano.

Responda SOMENTE um JSON, sem texto antes ou depois, no formato exato:
{"classificacoes": {"rotulo exato 1": "despesa", "rotulo exato 2": "receita", ...}}

As chaves do JSON devem ser IDÊNTICAS aos rótulos recebidos (mesma grafia,
maiúsculas/minúsculas, acentos)."""


def classificar_raizes_ambiguas_via_ia(raizes_ambiguas: list[dict], model: str | None = None) -> dict[str, str]:
    """Última etapa da extração de DRE por hierarquia (`extrair_hierarquia_dre`
    em dre_balancete_parser.py) — só as poucas linhas-raiz que sobraram sem
    sinal mecânico (prefixo "(-)"/"(+)"/"(=)", ou palavra óbvia de
    receita/despesa/resultado) chegam aqui. Na prática, testando 11 DREs
    reais em 25/08, isso é tipicamente 0-7 rótulos curtos por arquivo — não
    a planilha inteira. Prompt fixo, não depende do Supabase
    (`agent_prompts`) porque não é um dos agentes nomeados do pipeline, é
    um detalhe interno da extração.

    Retorna {rotulo: "despesa"|"receita"|"resultado"} — se a chamada falhar
    por qualquer motivo (rede, JSON malformado etc.), retorna {} e quem
    chamou decide o que fazer (as linhas continuam registradas como
    ambíguas, nunca descartadas silenciosamente)."""
    if not raizes_ambiguas:
        return {}
    rotulos = [r["rotulo"] for r in raizes_ambiguas]
    try:
        resultado = call_claude(
            _PROMPT_CLASSIFICACAO_RAIZES,
            {"rotulos_para_classificar": rotulos},
            model=model,
        )
        classificacoes = resultado.get("classificacoes", {})
        return {
            rotulo: tipo for rotulo, tipo in classificacoes.items()
            if rotulo in rotulos and tipo in ("despesa", "receita", "resultado")
        }
    except Exception as e:
        print(f"[extraction] classificação mínima de raízes ambíguas falhou (não bloqueia extração): {e}")
        return {}


def format_hierarquia_dre(hierarquia: dict, resultado_calculado: dict) -> str:
    """Formata o resultado de `extrair_hierarquia_dre` + `calcular_resultado_de_hierarquia`
    como texto compacto — usado quando a DRE não bate com `DRE_CATEGORIAS`
    (nomenclatura fora do padrão) e o caminho fino (`parse_dre_sheet`) não
    categorizou nada."""
    periodos_rotulos = hierarquia.get("periodos_rotulos") or []

    def _nomes_periodo(valores: dict) -> str:
        # "mes_01: 123" -> "2025 (ano completo): 123", quando a planilha
        # usa período nomeado livremente ou anual (ver periodos_rotulos);
        # pra série mensal de verdade, mantém "mes_01" (já é claro).
        partes = []
        for chave, v in valores.items():
            idx = int(chave.split("_")[1]) - 1
            nome = periodos_rotulos[idx] if idx < len(periodos_rotulos) and periodos_rotulos[idx] else chave
            partes.append(f"{nome}={v:,.2f}")
        return " | ".join(partes)

    lines = [
        "===== DRE (hierarquia extraída automaticamente por indentação/colunas, sem depender de nomenclatura fixa) =====",
        f"Modo de extração: {hierarquia['modo']} | Confiança: {'alta' if resultado_calculado.get('hierarquia_confiavel', True) else 'BAIXA — ver nota abaixo'}",
        "",
        "--- Linhas de RESULTADO já calculadas PELA PRÓPRIA planilha (mais confiáveis que qualquer soma nossa — "
        "use estas pra Margem Bruta/EBITDA/Líquida, calculando o % sobre a receita do MESMO período) ---",
    ]
    if resultado_calculado.get("linhas_resultado_da_fonte"):
        for item in resultado_calculado["linhas_resultado_da_fonte"]:
            lines.append(f"  {item['rotulo']}: {_nomes_periodo(item['valores'])}")
    else:
        lines.append("  (nenhuma linha de resultado pronta encontrada nesta planilha)")
    lines.append("")
    lines.append(
        "--- Aproximação AMPLA nossa (soma de tudo que classificamos como despesa, do custo direto "
        "até tributos — NÃO é Margem Bruta nem EBITDA específicos, só uma referência de sanidade) ---"
    )
    lines.append(f"Receita total (fonte: {resultado_calculado.get('receita_total_fonte', 'n/d')}): {resultado_calculado.get('receita_total')}")
    lines.append(f"Despesa total (todos os níveis somados): {resultado_calculado.get('despesa_total')}")
    lines.append(f"Margem ampla aproximada: {resultado_calculado.get('margem_operacional_pct')}%")
    if not resultado_calculado.get("hierarquia_confiavel", True):
        lines.append(
            "ATENÇÃO — confiança BAIXA: não foi detectada indentação real nessa planilha "
            "(todas as linhas no mesmo nível). Pode haver linhas-mãe (ex.: 'SG&A') somadas "
            "junto com seus próprios detalhes (ex.: 'Pessoal', 'Marketing'), duplicando o "
            "total. Trate os números acima como estimativa grosseira, não como EBITDA final."
        )
    lines.append("")
    lines.append("Todas as linhas classificadas por categoria (raiz | tipo | valor por período):")
    for rotulo, dados in hierarquia["raizes_classificadas"].items():
        lines.append(f"  {rotulo} | {dados['tipo']} | {_nomes_periodo(dados['valores'])}")
    return "\n".join(lines)


def excel_to_text(file_bytes: bytes, filename: str) -> tuple[str, list, dict, dict | None, dict | None, dict | None, dict | None]:
    """Converte um Excel em texto legível pro Claude — não interpreta,
    só descreve fielmente o que tem em cada aba, linha por linha.
    Testado localmente contra os 8 arquivos reais do BHub antes de ir
    pra produção (incluindo um com 24 abas e outro com erro #REF!).

    Retorna (texto, contas_balancete, dre_linhas, dre_hierarquia_info,
    mini_dre) — os 4 últimos já estruturados em código, prontos pra
    injetar em raw_extracted sem depender do LLM copiar de volta uma
    tabela grande. `dre_hierarquia_info` é None na maioria dos casos
    (quando `dre_linhas` já veio populada pelo caminho fino) — só vem
    preenchido quando a nomenclatura da DRE não bateu com
    `DRE_CATEGORIAS` e foi usado o fallback por indentação/colunas
    (achado real em 25/08). `mini_dre` é None quando não há DRE
    reconhecível nesse arquivo (cai pro EBITDA Bridge clássico).

    REVISADO em 20/08 contra um arquivo real (Nacional Controladoria) que
    tinha AS DUAS COISAS ao mesmo tempo: 12 abas mensais (MES 01..12) E
    uma aba "BALANCETE 2025" com a mesma informação já consolidada numa
    tabela só. Antes desta revisão, as duas eram enviadas — a mesma
    tabela de ~300 contas duplicada no prompt. Agora: se existe uma aba
    consolidada (qualquer nome — a detecção é pelo cabeçalho, não pelo
    nome da aba), ela é a fonte única e as abas mensais somem do despejo
    também. Só cai de volta pras 12 abas mensais se NENHUMA aba
    consolidada for encontrada."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return f"[ERRO ao abrir '{filename}' como Excel: {e}]", [], {}

    parts = [f"===== ARQUIVO: {filename} ====="]
    skip_sheets = set()
    contas: list = []

    consolidado = detect_consolidated_balancete(wb)
    if consolidado:
        contas = parse_consolidated_balancete(wb, consolidado)
        skip_sheets.add(consolidado["aba"])
        # As 12 abas mensais (se existirem) ficam redundantes com a
        # consolidada — não processa as duas, e tira ambas do despejo cru.
        monthly_sheets = detect_monthly_sheets(wb)
        skip_sheets |= {name for _, name in monthly_sheets}
        parts.append(
            f"===== SÉRIE MENSAL CONSOLIDADA (extraída automaticamente da aba '{consolidado['aba']}') =====\n"
            f"{len(contas)} contas encontradas — não precisa copiar no seu output, "
            "o código já anexa em raw_extracted.series_contabeis.\n"
            "ATENÇÃO: contas de resultado (receita/despesa) costumam vir ACUMULADAS no "
            "ano-calendário — para o valor do mês isolado, calcule a diferença entre um "
            "mês e o anterior. Contas de balanço (ativo/passivo) já são o saldo no fim "
            "daquele mês, sem precisar de ajuste."
        )
    else:
        monthly_sheets = detect_monthly_sheets(wb)
        skip_sheets |= {name for _, name in monthly_sheets}
        if monthly_sheets:
            accounts_dict = merge_monthly_balancete(wb, monthly_sheets)
            contas = [
                {"conta": nome, "codigo": codigo, "tipo": data["tipo"],
                 "valores": {f"mes_{m:02d}": v for m, v in data["meses"].items()}}
                for (nome, codigo), data in accounts_dict.items()
            ]
            parts.append(format_merged_table(accounts_dict))

    dre_linhas: dict = {}
    dre_hierarquia_info: dict | None = None
    dre_deteccao = detect_dre_sheet(wb)
    multi_entidade_ambigua_nota = None
    if dre_deteccao and dre_deteccao.get("multi_entidade_ambigua"):
        # Achado real em 27/08 (deal Irko): a aba tem 2+ blocos de
        # empresa empilhados (holding), mas nenhum bloco "combinado"/
        # "consolidado" claro pra escolher sozinho. Adivinhar aqui é
        # exatamente o bug que gerou o resultado errado do Irko (pegar
        # uma subsidiária como se fosse o grupo todo) — então NÃO
        # tenta extração fina nem hierarquia neste arquivo. Trata como
        # "sem DRE reconhecível" (dre_deteccao vira None daqui pra
        # frente) e deixa uma nota explícita, que vira red flag visível
        # pro Thiago/parceiro resolver manualmente, em vez de um número
        # errado sem aviso nenhum.
        multi_entidade_ambigua_nota = (
            f"Aba '{dre_deteccao['aba']}' tem {len(dre_deteccao['blocos_detectados'])} blocos de "
            f"empresa/entidade ({', '.join(dre_deteccao['blocos_detectados'])}), mas nenhum bloco "
            "claramente 'combinado'/'consolidado' foi identificado — extração financeira NÃO "
            "prosseguiu neste arquivo pra evitar pegar uma subsidiária isolada como se fosse o "
            "total. Confirmar com o parceiro qual bloco (ou soma) representa o deal, ou renomear "
            "o bloco correto pra incluir 'combinado'/'consolidado'/'total grupo' no rótulo."
        )
        parts.append(
            f"===== AVISO: MÚLTIPLAS EMPRESAS DETECTADAS NA ABA '{dre_deteccao['aba']}' =====\n"
            + multi_entidade_ambigua_nota
        )
        dre_hierarquia_info = {"multi_entidade_ambigua": True, "nota": multi_entidade_ambigua_nota,
                                "blocos_detectados": dre_deteccao["blocos_detectados"]}
        dre_deteccao = None
    if dre_deteccao:
        dre_linhas = parse_dre_sheet(wb, dre_deteccao)
        skip_sheets.add(dre_deteccao["aba"])
        # ATENÇÃO (bug real achado em 25/08): `parse_dre_sheet` NUNCA
        # retorna vazio quando há linhas com valor — quando o rótulo não
        # bate nenhuma categoria conhecida, ele usa o RÓTULO BRUTO como
        # chave (`categoria or rotulo_str`), em vez de descartar a linha.
        # `if dre_linhas:` sozinho, portanto, é sempre True e nunca cai no
        # fallback novo.
        #
        # SEGUNDO BUG (achado testando os 11 arquivos reais de ponta a
        # ponta): contar "quantas categorias bateram" também não basta —
        # em 3 dos 11 arquivos reais, entre 2 e 3 categorias bateram, mas
        # NENHUMA das combinações que `calcular_ebitda_de_dre` realmente
        # usa (receita_liquida + despesas_operacionais_total, OU
        # resultado_liquido) — o EBITDA saía TODO None mesmo "tendo
        # categorizado algo". O critério certo é perguntar diretamente:
        # essa categorização vai produzir um EBITDA de verdade?
        #
        # TERCEIRO BUG (achado real no Plannea, 26/08): mesmo esse
        # critério "direto" não bastou — `calcular_ebitda_de_dre` usa só
        # as categorias FIXAS antigas (`despesas_operacionais_total`
        # etc.), que essa DRE não bate ("CUSTO OPERACIONAL" +
        # "DESPESAS ADMINISTRATIVAS" separados, não uma categoria única
        # reconhecida) — mesmo já tendo tudo que `montar_mini_dre`
        # precisa (via busca mais solta por palavra-chave). Sem essa
        # segunda checagem, o caminho fino era descartado por engano e o
        # pipeline caía no fallback de hierarquia, que SOMA TODOS OS
        # PERÍODOS numa métrica só — nessa DRE, 18 meses (jan/2025 a
        # jun/2026) somados viraram uma "receita" de R$42,56M sem
        # sentido nenhum, e um EBITDA implausível de R$8,89M reportado
        # ao usuário. Agora testa os dois: categorias fixas OU
        # mini-DRE por palavra-chave — qualquer um que funcione já
        # confirma que o caminho fino é utilizável.
        ebitda_teste = calcular_ebitda_de_dre(dre_linhas)
        periodos_rotulos_teste = _rotulos_legiveis_periodo(dre_deteccao["meses_para_coluna"])
        mini_dre_teste = montar_mini_dre(dre_linhas, None, periodos_rotulos_teste)
        caminho_fino_produz_ebitda = (
            ebitda_teste.get("ebitda_bottom_up_receita_menos_despesas") is not None
            or ebitda_teste.get("ebitda_top_down_lucro_liquido") is not None
            or bool(mini_dre_teste)
        )
        if caminho_fino_produz_ebitda:
            periodos_rotulos_fino = periodos_rotulos_teste
            dre_linhas_para_prompt, periodos_rotulos_para_prompt = dre_linhas, periodos_rotulos_fino
            if dre_deteccao["granularidade"] == "mensal":
                # Agrupamento trimestral SÓ NO TEXTO QUE VAI PRO AGENTE
                # (26/08, achado real — deal "Nacional", 12 meses de
                # DRE): sem isso, o contexto ficou grande o bastante pra
                # resposta do agente `opinion` ser truncada pelo limite
                # de tokens de saída do modelo (job quebrou depois de já
                # ter sido gerado e pago). NÃO agrupa `dre_linhas` em si
                # — essa variável segue alimentando `mini_dre`/
                # `tabela_viabilidade_financeira`/Excel, que devem manter
                # o detalhe mensal completo (o Thiago pediu agrupamento
                # só "no PPT", que é tratado separadamente em
                # generate_outputs.py). Só ativa com mais de 6 períodos.
                dre_linhas_para_prompt, periodos_rotulos_para_prompt = agrupar_dre_linhas_por_trimestre(dre_linhas, periodos_rotulos_fino)
            parts.append(format_dre_table(dre_linhas_para_prompt, periodos_rotulos_para_prompt, dre_deteccao["granularidade"]))
        else:
            # Caminho fino (regex de nomenclatura conhecida) não achou
            # nada — achado real em 25/08: acontece sempre que a empresa
            # usa nomenclatura fora do padrão formal ("RECEITAS" em vez
            # de "Receita Líquida", por exemplo). Fallback: hierarquia
            # por indentação/colunas (não depende de nomenclatura), com
            # uma chamadinha de IA só pras poucas linhas-raiz que
            # sobrarem sem sinal mecânico (tipicamente 0-7 rótulos
            # curtos, nunca a planilha inteira).
            # Zera dre_linhas (mesmo tendo rótulos brutos como chave) —
            # sem isso, `dre_estruturada` chegaria não-vazia no main() e
            # o fallback novo nunca seria escolhido lá (prioridade é dela).
            dre_linhas = {}
            hierarquia = extrair_hierarquia_dre(wb, dre_deteccao)
            if hierarquia["raizes_classificadas"] or hierarquia["raizes_ambiguas"]:
                classificacoes_ia = classificar_raizes_ambiguas_via_ia(hierarquia["raizes_ambiguas"])
                ainda_ambiguas = []
                for r in hierarquia["raizes_ambiguas"]:
                    tipo_ia = classificacoes_ia.get(r["rotulo"])
                    if tipo_ia:
                        tipo_final = "resultado_final" if tipo_ia == "resultado" else tipo_ia
                        hierarquia["raizes_classificadas"][r["rotulo"]] = {
                            "tipo": tipo_final, "valores": r["valores"], "detalhes": r["detalhes"],
                        }
                    else:
                        ainda_ambiguas.append(r)
                hierarquia["raizes_ambiguas"] = ainda_ambiguas

                resultado_calc = calcular_resultado_de_hierarquia(hierarquia)
                dre_hierarquia_info = {"hierarquia": hierarquia, "resultado": resultado_calc}
                parts.append(format_hierarquia_dre(hierarquia, resultado_calc))

    # Mini-DRE por período (achado real em 26/08, pedido do Thiago: mini
    # DRE visível no PPT/Excel, estilo teaser de M&A) — determinística,
    # roda nos 2 caminhos (fino e hierarquia), usando os rótulos de
    # período reais ("2025 (Ano Completo)", não "mes_01" genérico).
    mini_dre = None
    if dre_deteccao:
        periodos_rotulos = _rotulos_legiveis_periodo(dre_deteccao["meses_para_coluna"])
        hierarquia_para_mini_dre = dre_hierarquia_info["hierarquia"] if dre_hierarquia_info else None
        mini_dre = montar_mini_dre(dre_linhas or None, hierarquia_para_mini_dre, periodos_rotulos)

    # Tabela "Viabilidade Financeira" (26/08, pedido do Thiago) — mesma
    # ideia da mini-DRE, estendida com D&A/Despesas Gerais/Lucro
    # Operacional/Margem EBITDA, no formato exato que ele validou.
    tabela_viabilidade = None
    if dre_deteccao:
        tabela_viabilidade = montar_tabela_viabilidade_financeira(dre_linhas or None, hierarquia_para_mini_dre, periodos_rotulos)

    # RBT12 validado + MRR estimado (achado real em 26/08, pedido do
    # Thiago: "coloque MRR" — RBT12 é conceito TRIBUTÁRIO específico
    # (Anexo III do Simples Nacional, definido por lei como receita bruta
    # dos últimos 12 MESES) e continua necessário internamente pro
    # cálculo de alíquota — mas reportar "RBT12" pro usuário como se
    # fosse o KPI de porte do negócio confunde mais do que ajuda,
    # principalmente numa empresa de receita recorrente (BPO), onde MRR
    # é a métrica padrão do setor.
    #
    # BUG REAL corrigido junto (achado revisando este mesmo trecho): o
    # cálculo antigo de `rbt12_real` somava TODOS os valores disponíveis
    # sem checar se eram realmente 12 meses — numa DRE com granularidade
    # "periodo_livre" (ex.: 1 ano + 1 trimestre, como no BPO Innova),
    # isso somava ~15 meses de receita como se fossem 12, inflando o
    # RBT12 usado no cálculo de alíquota tributária. Agora só soma como
    # RBT12 quando a granularidade é realmente mensal.
    kpis_periodo = None
    if dre_deteccao:
        granularidade = dre_deteccao["granularidade"]
        receita_bruta_periodos = dre_linhas.get("receita_bruta") if dre_linhas else None
        if receita_bruta_periodos is None and mini_dre and mini_dre.get("linhas"):
            receita_bruta_periodos = {f"p{i:02d}": l["receita_bruta"] for i, l in enumerate(mini_dre["linhas"])}

        rbt12_dre_valido, mrr_estimado, mrr_fonte = None, None, None
        if receita_bruta_periodos:
            valores = [v for v in receita_bruta_periodos.values() if isinstance(v, (int, float))]
            if granularidade == "mensal" and len(valores) >= 11:
                rbt12_dre_valido = sum(valores[-12:])
                mrr_estimado, mrr_fonte = valores[-1], "dre_ultimo_mes"
            elif granularidade == "anual" and valores:
                mrr_estimado, mrr_fonte = valores[-1] / 12, "dre_ano_mais_recente_dividido_12"
            # granularidade "periodo_livre" (ex.: ano + trimestre juntos):
            # duração de cada período não é padronizada o bastante pra
            # estimar MRR com segurança aqui — cai pro faturamento_mensal
            # do formulário como fonte, tratado abaixo.
        kpis_periodo = {
            "rbt12_dre_valido": round(rbt12_dre_valido, 2) if rbt12_dre_valido else None,
            "mrr_estimado_dre": round(mrr_estimado, 2) if mrr_estimado else None,
            "mrr_fonte": mrr_fonte,
            "granularidade_dre": granularidade,
        }

    # Se já tiramos dado financeiro estruturado (DRE ou Balancete) deste
    # arquivo, as abas que sobram (Menu, CMV, Despesas, R.F., DFC,
    # Inventário, etc.) hoje não alimentam nada no pipeline — o
    # financial_analysis só lê dre_estruturada/series_contabeis. Confirmado
    # com o Thiago em 21/08: cortar o despejo cru delas, custo sem retorno.
    # Se algum dia precisarmos (análise de pessoal/fluxo de caixa detalhado),
    # a resposta certa é escrever um parser dedicado pra elas, não reativar
    # o despejo cru.
    houve_extracao_estruturada = bool(consolidado) or bool(dre_deteccao)
    sheet_names = [] if houve_extracao_estruturada else [s for s in wb.sheetnames if s not in skip_sheets][:MAX_SHEETS]
    abas_nao_incluidas = [s for s in wb.sheetnames if s not in skip_sheets] if houve_extracao_estruturada else []
    if abas_nao_incluidas:
        parts.append(
            f"(outras {len(abas_nao_incluidas)} abas deste arquivo não incluídas — "
            f"{', '.join(abas_nao_incluidas[:8])}{'...' if len(abas_nao_incluidas) > 8 else ''} — "
            "não alimentam o pipeline hoje; avisar se precisar do conteúdo delas.)"
        )
    if len(wb.sheetnames) - len(skip_sheets) > MAX_SHEETS and not houve_extracao_estruturada:
        parts.append(f"(arquivo tem mais abas — mostrando as primeiras {MAX_SHEETS} além da série mensal/DRE)")

    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        parts.append(f"\n--- Aba: {sheet_name} ---")
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            if row_count >= MAX_ROWS_PER_SHEET:
                parts.append(f"[... aba truncada em {MAX_ROWS_PER_SHEET} linhas ...]")
                break
            row_trimmed = row[:MAX_COLS_PER_SHEET]
            if all(v is None for v in row_trimmed):
                row_count += 1
                continue
            values = [("" if v is None else str(v)) for v in row_trimmed]
            parts.append(f"L{row_count+1}: " + " | ".join(values))
            row_count += 1

    return "\n".join(parts), contas, dre_linhas, dre_hierarquia_info, mini_dre, kpis_periodo, tabela_viabilidade


def select_relevant_accounts(series: list, top_n: int = MAX_ACCOUNTS_FOR_FINANCIAL_AGENT) -> list:
    """Reduz a série contábil consolidada (potencialmente 300+ contas) às
    que interessam pra análise financeira: maior volatilidade mês a mês
    (candidata a anomalia) combinada com materialidade absoluta. Isso é
    estatística determinística — não pedimos pro modelo vasculhar a
    tabela inteira pra achar o que o código já sabe calcular em O(n)."""
    def score(item: dict) -> float:
        valores = [v for v in item.get("valores", {}).values() if isinstance(v, (int, float))]
        if not valores:
            return 0.0
        materialidade = max(abs(v) for v in valores)
        if len(valores) < 3:
            return materialidade
        deltas = [b - a for a, b in zip(valores, valores[1:])]
        volatilidade = statistics.pstdev(deltas)
        # combina os dois sinais — uma conta pequena mas muito instável
        # (ex.: pró-labore que dobra num mês) também precisa aparecer,
        # não só as maiores em valor absoluto.
        return volatilidade + 0.1 * materialidade

    return sorted(series, key=score, reverse=True)[:top_n]


def supabase_request(method: str, path: str, body: dict | None = None) -> dict:
    """Chamada simples à API REST do Supabase (PostgREST), usando a service_role key."""
    url = os.environ["SUPABASE_URL"].rstrip("/") + "/rest/v1/" + path
    key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=representation",
    }
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        print(f"Erro Supabase [{method} {path}]: {e.code} {e.read().decode()}", file=sys.stderr)
        raise


def get_deal_data(deal_id: str) -> dict:
    rows = supabase_request("GET", f"deal_data?deal_id=eq.{deal_id}&order=created_at.desc&limit=1")
    if not rows:
        raise SystemExit(f"Nenhum deal_data encontrado para o deal {deal_id} — rode o Agent 00 primeiro.")
    return rows[0]


def download_from_storage(storage_ref: str) -> bytes:
    url = os.environ["SUPABASE_URL"].rstrip("/") + "/storage/v1/object/deal-files/" + storage_ref
    key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
    req = urllib.request.Request(url, headers={"apikey": key, "Authorization": f"Bearer {key}"})
    with urllib.request.urlopen(req) as resp:
        return resp.read()


def run_extraction(deal_id: str):
    """Caminho próprio do Agent 00: baixa os arquivos do deal, converte pra
    texto, chama o Claude, e CRIA um deal_data novo (os outros agentes só
    leem um deal_data que já existe — este é quem produz).

    A série mensal consolidada (quando existe) é calculada pelo código e
    injetada direto no resultado — não pedimos pro Claude copiar de volta
    uma tabela de centenas de contas x 12 meses que o código já tem. Isso
    era o que estava estourando o timeout de 10 minutos (visto em 19/08)."""
    inicio = datetime.now(timezone.utc)  # achado real em 28/08 (item 2 da lista
    # de melhorias): sem isso, o site não tinha como saber quanto tempo CADA
    # agente levou de verdade — só "há quanto tempo terminou". Grava junto
    # com o resultado (`started_at`), pro front calcular a duração real de
    # cada estágio sem precisar aproximar assumindo que todos começam juntos.
    files = supabase_request("GET", f"files?deal_id=eq.{deal_id}")
    if not files:
        raise SystemExit(f"Nenhum arquivo encontrado para o deal {deal_id} — nada para extrair.")

    agent_version = get_active_agent_version("extraction")

    combined_text = []
    raw_bytes_for_checksum = b""
    code_computed_series = []  # preenchido diretamente pelo código, não pelo LLM
    code_computed_dre = {}     # idem, pela mesma razão
    code_computed_dre_hierarquia = {}  # idem — fallback quando a DRE não bate nomenclatura conhecida
    code_computed_mini_dre = {}  # idem — DRE simplificada por período, pro PPT/Excel
    code_computed_tabela_viabilidade = {}  # idem — tabela "Viabilidade Financeira" estendida (D&A, Margem EBITDA), pro Excel
    code_computed_kpis_periodo = {}  # idem — RBT12 validado + MRR estimado, por arquivo
    code_computed_formulario_raw = None  # idem — respostas canônicas do formulário, se reconhecido
    for f in files:
        try:
            content = download_from_storage(f["storage_ref"])
        except urllib.error.HTTPError as e:
            combined_text.append(f"===== ARQUIVO: {f['original_filename']} =====\n[ERRO ao baixar do Storage: {e.code}]")
            continue
        raw_bytes_for_checksum += content
        filename = f["original_filename"] or f["storage_ref"]

        # Formulário do parceiro: cabeçalho fixo, reconhecido por palavra-
        # chave (ver formulario_mapper.py). Se reconhecer, processa 100% em
        # código e NUNCA manda o conteúdo cru pro Claude — é a maior fatia
        # de redução de token desta rodada (confirmado com o Thiago em
        # 21/08: "o impacto agora em tokens deve ser gigantesco").
        formulario_detectado = detectar_e_extrair_formulario(content)
        if formulario_detectado is not None:
            code_computed_formulario_raw = formulario_detectado
            combined_text.append(
                f"===== ARQUIVO: {filename} =====\n"
                f"(Formulário do parceiro reconhecido e processado 100% em código — "
                f"{len(formulario_detectado)} campos extraídos por palavra-chave, sem IA. "
                "Conteúdo não incluído aqui de propósito, pra não gastar token com o que "
                "o código já resolveu sozinho.)"
            )
            continue

        texto, contas, dre, dre_hierarquia_info, mini_dre, kpis_periodo, tabela_viabilidade = excel_to_text(content, filename)
        combined_text.append(texto)
        for conta in contas:
            conta.setdefault("arquivo", filename)
            code_computed_series.append(conta)
        if dre:
            code_computed_dre[filename] = dre
        if dre_hierarquia_info:
            code_computed_dre_hierarquia[filename] = dre_hierarquia_info
        if mini_dre:
            code_computed_mini_dre[filename] = mini_dre
        if tabela_viabilidade:
            code_computed_tabela_viabilidade[filename] = tabela_viabilidade
        if kpis_periodo:
            code_computed_kpis_periodo[filename] = kpis_periodo

    full_dump = "\n\n".join(combined_text)
    checksum = hashlib.sha256(raw_bytes_for_checksum).hexdigest()

    existing = supabase_request("GET", f"deal_data?deal_id=eq.{deal_id}&checksum=eq.{checksum}")
    if existing:
        print("[extraction] mesmo conjunto de arquivos já extraído antes — pulando (idempotência).")
        return

    output = call_claude(
        agent_version["system_prompt"], {"arquivos_em_texto": full_dump},
        model=agent_version.get("model"),
    )

    # Injeta a série e a DRE calculadas pelo código — sobrepõe o que o LLM
    # eventualmente tenha tentado copiar (nunca confiamos numa cópia manual
    # de centenas de linhas, e a DRE já vem pronta e categorizada).
    if code_computed_series:
        output.setdefault("raw_extracted", {})["series_contabeis"] = code_computed_series
    if code_computed_dre:
        output.setdefault("raw_extracted", {})["dre_estruturada"] = code_computed_dre
    if code_computed_dre_hierarquia:
        # Fallback (25/08) — só populado quando a DRE não bateu nenhuma
        # categoria de `DRE_CATEGORIAS` (nomenclatura fora do padrão
        # formal). Formato diferente de `dre_estruturada`: aqui é
        # {arquivo: {"hierarquia": {...}, "resultado": {...}}}, não
        # {arquivo: {categoria: {mes: valor}}} — quem consome (ver
        # main(), branch financial_analysis) precisa checar os dois.
        output.setdefault("raw_extracted", {})["dre_hierarquia_aproximada"] = code_computed_dre_hierarquia
    if code_computed_mini_dre:
        # DRE simplificada por período (26/08, pedido do Thiago pra ter
        # uma mini-DRE visível no PPT/Excel, estilo teaser de M&A) —
        # {arquivo: {"linhas": [...], "fonte": "fino"|"hierarquia"}}.
        output.setdefault("raw_extracted", {})["mini_dre"] = code_computed_mini_dre
    if code_computed_tabela_viabilidade:
        # Tabela "Viabilidade Financeira" completa (26/08, pedido do
        # Thiago — validado célula por célula contra o exemplo dele) —
        # {arquivo: {"linhas": [...], "fonte": ..., "d_a_reconhecido": bool}}.
        output.setdefault("raw_extracted", {})["tabela_viabilidade_financeira"] = code_computed_tabela_viabilidade

    if code_computed_formulario_raw is not None:
        # RBT12 real, a partir da receita bruta de 12 meses já extraída da
        # DRE — muito mais confiável que aproximar por faturamento_mensal
        # × 12 (decisão do Thiago em 21/08: "pegaremos na DRE"). Só cai pra
        # aproximação se não existir DRE nesse deal.
        # BUG REAL CORRIGIDO EM 26/08: a versão antiga somava TODOS os
        # valores de receita disponíveis sem checar se eram realmente 12
        # meses — numa DRE "período livre" (ex.: 1 ano + 1 trimestre),
        # isso somava ~15 meses como se fossem 12. Agora usa
        # `kpis_periodo`, que só considera RBT12 válido quando a
        # granularidade é mensal de verdade.
        primeiro_kpis_periodo = next(iter(code_computed_kpis_periodo.values()), {}) if code_computed_kpis_periodo else {}
        rbt12_real = primeiro_kpis_periodo.get("rbt12_dre_valido")

        mapeado = mapear_formulario(code_computed_formulario_raw, rbt12_real=rbt12_real)

        # MRR (Receita Mensal Recorrente) — achado real em 26/08, pedido
        # do Thiago: "coloque MRR" no lugar de RBT12 como KPI de negócio
        # reportado ao usuário. RBT12 continua existindo (acima) só pro
        # cálculo INTERNO de alíquota do Simples Nacional — não é o KPI
        # certo pra descrever o porte de um negócio de receita recorrente
        # (BPO). Prioridade de fonte: DRE mensal (mais objetivo) > DRE
        # anual dividido por 12 (aproximação) > faturamento_mensal do
        # formulário (menos objetivo, preenchido à mão, mas sempre
        # disponível como último recurso).
        mrr_estimado = primeiro_kpis_periodo.get("mrr_estimado_dre")
        mrr_fonte = primeiro_kpis_periodo.get("mrr_fonte")
        if mrr_estimado is None and mapeado.get("faturamento_mensal") is not None:
            mrr_estimado = mapeado["faturamento_mensal"]
            mrr_fonte = "formulario_faturamento_mensal"
        mapeado["mrr_estimado"] = round(mrr_estimado, 2) if mrr_estimado is not None else None
        mapeado["mrr_fonte"] = mrr_fonte
        mapeado["nota_sobre_kpis_financeiros"] = (
            "Para descrever o PORTE/RECEITA do negócio ao usuário (sumário executivo, "
            "parecer, KPIs de valuation), use mrr_estimado (Receita Mensal Recorrente) — "
            "é o padrão do setor pra negócios de receita recorrente (BPO/serviços). "
            "rbt12 existe só pro cálculo INTERNO de alíquota do Simples Nacional (é um "
            "conceito tributário, definido por lei sobre 12 meses fechados) — não é KPI "
            "de negócio e não deve ser citado como tal na narrativa."
        )

        # Bloco A e B já calculados aqui, na extração — Complexity e
        # Viabilidade Financeira (agentes separados) só vão COPIAR isto
        # pra agent_runs depois, sem recalcular nada.
        #
        # PRIORIDADE DE FONTE PRA MARGEM BRUTA/CUSTO FOLHA (decisão
        # explícita do Thiago em 26/08: "quem manda é a DRE, não o
        # formulário"): testando contra o BPO Innova real, o formulário
        # dava margem bruta de 46,89% (faturamento/folha/sistemas do
        # formulário não batem com nenhum período real da DRE — parecem
        # um "instantâneo" mais recente, preenchido à mão, não uma
        # extração da DRE), enquanto a DRE bate quase exato com o que o
        # time humano já tinha validado (55%/58%). Usa a DRE (via
        # `extrair_margem_bruta_de_dre`, busca por palavra-chave nos
        # rótulos — funciona tanto no caminho fino quanto na hierarquia)
        # quando ela existir E tiver os componentes mínimos (receita,
        # despesa com pessoal, custo sistemas); só cai pro formulário
        # quando a DRE não tem esses componentes reconhecíveis.
        primeira_dre_fina = next(iter(code_computed_dre.values()), None) if code_computed_dre else None
        primeira_hierarquia = next(iter(code_computed_dre_hierarquia.values()), {}).get("hierarquia") if code_computed_dre_hierarquia else None
        margem_de_dre = extrair_margem_bruta_de_dre(primeira_dre_fina, primeira_hierarquia, mapeado["regime_tributario"])

        if margem_de_dre:
            # Período mais recente = último do dict (mes_01=2025, mes_02=2026
            # neste deal — a ordem já vem cronológica de `meses_para_coluna`).
            ultimo_periodo = list(margem_de_dre["margem_bruta_pct_por_periodo"])[-1]
            margem_bruta_pct_final = margem_de_dre["margem_bruta_pct_por_periodo"][ultimo_periodo]
            custo_folha_pct = margem_de_dre["custo_folha_pct_por_periodo"][ultimo_periodo]
            margem_bruta_calculada = {
                "margem_bruta_pct": margem_bruta_pct_final,
                "fonte": f"dre_{margem_de_dre['fonte']}",
                "motivo": (
                    "DRE disponível e com os componentes necessários (Receita, Despesa com "
                    "Pessoal, Custo Sistemas) reconhecidos — prioridade sobre o formulário "
                    "por decisão do Thiago em 26/08 (o formulário reflete um instantâneo "
                    "recente preenchido à mão, a DRE é o histórico contábil real)."
                ),
                "todos_periodos": margem_de_dre["margem_bruta_pct_por_periodo"],
            }
        else:
            # BUG REAL CORRIGIDO EM 26/08: esta variável só existia dentro
            # deste branch, mas era referenciada mais abaixo (raw_extracted)
            # de forma incondicional — quebrava com NameError toda vez que a
            # DRE fosse a fonte usada (branch acima), que passou a ser o
            # caminho mais comum depois da decisão "quem manda é a DRE".
            margem_formulario = calcular_margem_bruta(
                mapeado["faturamento_mensal"], mapeado["folha_informada"], mapeado["custo_sistemas"],
                mapeado["regime_tributario"], mapeado["rbt12"],
            )
            margem_bruta_pct_final = margem_formulario["margem_bruta_pct"]
            custo_folha_pct = (
                100 * mapeado["folha_informada"] / mapeado["faturamento_mensal"]
                if mapeado["faturamento_mensal"] else None
            )
            motivo_formulario = (
                "Nenhuma DRE disponível nesse deal."
                if not (code_computed_dre or code_computed_dre_hierarquia)
                else "DRE disponível, mas sem os componentes necessários reconhecíveis "
                     "(Receita/Despesa com Pessoal/Custo Sistemas) — caiu pro formulário."
            )
            margem_bruta_calculada = {
                **margem_formulario,
                "fonte": "formulario",
                "motivo": motivo_formulario,
            }

        bloco_a = avaliar_viabilidade_financeira(
            margem_bruta_pct_final, custo_folha_pct,
            mapeado["inadimplencia_media_pct"], mapeado["churn_medio_pct"], mapeado["perfil_restrito_presente"],
        )
        bloco_a["margem_bruta_fonte"] = margem_bruta_calculada["fonte"]
        bloco_a["margem_bruta_motivo"] = margem_bruta_calculada["motivo"]

        # Achado real em 26/08 (pedido explícito do Thiago): "sempre que
        # houver DRE, ignorar o formulário completamente" pra números
        # FINANCEIROS (receita, margem, custo folha, custo sistemas) —
        # não pra dado que só existe no formulário mesmo (churn,
        # concentração, número de clientes/colaboradores, outsourcing,
        # regime tributário — a DRE não informa nada disso). Reforça a
        # nota já criada acima com essa regra explícita, agora que já
        # sabemos se a DRE de fato foi usada (`margem_de_dre` truthy).
        if margem_de_dre:
            mapeado["nota_sobre_kpis_financeiros"] += (
                " IMPORTANTE: a DRE deste deal tem os dados financeiros necessários e "
                "JÁ FOI USADA como fonte (ver margem_bruta_fonte/margem_bruta_calculada) "
                "— para receita, margem, custo de folha e custo de sistemas, ignore "
                "completamente os valores equivalentes do formulário "
                "(faturamento_mensal, folha_informada, custo_sistemas) mesmo que "
                "divirjam da DRE; a DRE é a fonte de verdade neste deal. O formulário "
                "continua sendo a única fonte pra dado que a DRE não informa (churn, "
                "concentração, número de clientes/colaboradores, outsourcing, regime "
                "tributário, sistemas utilizados)."
            )

        # Custo Sistemas % (Bloco B) — mesma prioridade DRE > formulário
        # (decisão do Thiago em 26/08: "sempre que houver DRE, ignorar o
        # formulário completamente"). Antes só a Margem Bruta seguia essa
        # regra; o Bloco B continuava lendo direto do formulário mesmo
        # quando a DRE já tinha o dado — reaproveita o mesmo
        # `margem_de_dre` calculado acima (já busca Custo Sistemas por
        # palavra-chave), sem nova chamada.
        if margem_de_dre and margem_de_dre.get("custo_sistemas_pct_por_periodo"):
            ultimo_periodo_b = list(margem_de_dre["margem_bruta_pct_por_periodo"])[-1]
            custo_sistemas_pct_dre = margem_de_dre["custo_sistemas_pct_por_periodo"].get(ultimo_periodo_b)
        else:
            custo_sistemas_pct_dre = None
        if custo_sistemas_pct_dre is not None:
            custo_sistemas_pct = custo_sistemas_pct_dre
        else:
            custo_sistemas_pct = (
                100 * mapeado["custo_sistemas"] / mapeado["faturamento_mensal"]
                if mapeado["faturamento_mensal"] else None
            )
        bloco_b = avaliar_complexidade_operacional(
            custo_sistemas_pct, mapeado["localizacao_fora_grande_sp"],
            mapeado["numero_clientes"], mapeado["numero_colaboradores"],
            mapeado["concentracao_top10_pct"], mapeado["segmentos_sensiveis_presentes"],
            mapeado["sistema_financeiro_e_omie"], mapeado["sistema_utilizado_texto"],
            mapeado["sistema_hospedagem"], mapeado["outsourcing_pessoas_pct"], mapeado["outsourcing_sistemas_pct"],
            mapeado.get("outsourcing_pessoas_pct_faturamento"), mapeado.get("outsourcing_sistemas_pct_faturamento"),
        )

        output.setdefault("structured", {}).update(
            {k: v for k, v in mapeado.items() if k != "avisos_mapeamento"}
        )
        output.setdefault("raw_extracted", {})["margem_bruta_calculada"] = margem_bruta_calculada
        output.setdefault("raw_extracted", {})["viabilidade_financeira_calculada"] = bloco_a
        output.setdefault("raw_extracted", {})["complexidade_operacional_calculada"] = bloco_b.to_agent_run_output()
        output.setdefault("raw_extracted", {})["riscos_operacionais_calculados"] = avaliar_riscos_operacionais(mapeado)
        output.setdefault("raw_extracted", {})["riscos_integracao_calculados"] = avaliar_riscos_integracao(mapeado)
        # Red flag de ERP pouco difundido — código, não IA (26/08, pedido
        # do Thiago: "é alto pois não está na nossa lista e nós
        # identificamos se esse ERP é pouco utilizado ou não"). None
        # quando o ERP é reconhecido ou não informado — nesse caso não
        # salva nada (evita raw_extracted poluído com chave vazia à toa).
        red_flag_erp = gerar_red_flag_erp(mapeado["sistema_utilizado_texto"], mapeado["sistema_hospedagem"])
        if red_flag_erp:
            output.setdefault("raw_extracted", {})["red_flag_erp_calculado"] = red_flag_erp
        if mapeado["avisos_mapeamento"]:
            output.setdefault("avisos", [])
            output["avisos"].extend(mapeado["avisos_mapeamento"])

    deal_data_row = supabase_request("POST", "deal_data", {
        "deal_id": deal_id,
        "schema_version": "v1",
        "structured": output.get("structured", {}),
        "raw_extracted": output.get("raw_extracted", {}),
        "checksum": checksum,
    })
    deal_data_id = deal_data_row[0]["id"] if isinstance(deal_data_row, list) else deal_data_row["id"]

    # Registra também em agent_runs, pro AI Audit Log não ter um buraco
    # justamente no primeiro agente da cadeia.
    supabase_request("POST", "agent_runs", {
        "deal_id": deal_id,
        "agent_version_id": agent_version["id"],
        "deal_data_id": deal_data_id,
        "input_hash": checksum,
        "status": "completed",
        "output": output,
        "confidence": output.get("confidence"),
        "started_at": inicio.isoformat(),
    })
    print(f"[extraction] deal_data criado ({deal_data_id}), confidence={output.get('confidence')}")


def get_active_agent_version(agent_name: str) -> dict:
    agents = supabase_request("GET", f"agents?name=eq.{agent_name}")
    if not agents:
        raise SystemExit(f"Agente '{agent_name}' não encontrado na tabela agents.")
    agent_id = agents[0]["id"]

    versions = supabase_request(
        "GET", f"agent_versions?agent_id=eq.{agent_id}&active=eq.true&order=created_at.desc&limit=1"
    )
    if not versions:
        raise SystemExit(f"Nenhuma versão ativa para o agente '{agent_name}'.")
    version = versions[0]

    prompts = supabase_request("GET", f"agent_prompts?agent_version_id=eq.{version['id']}")
    if not prompts:
        raise SystemExit(f"Sem system_prompt gravado para a versão {version['id']} de '{agent_name}'.")

    version["system_prompt"] = prompts[0]["system_prompt"]
    return version


def call_claude(system_prompt: str, deal_data: dict, other_outputs: dict | None = None,
                 model: str | None = None) -> dict:
    """Chama o Claude Code em modo headless e valida que a resposta é JSON.

    O prompt vai por stdin, não como argumento de linha de comando — prompts
    grandes (planilhas reais podem passar de 100 mil caracteres) estouram o
    limite do sistema operacional para argumentos (visto na prática em
    19/08: OSError "Argument list too long"). Documentação oficial confirma
    stdin como o caminho certo para isso: `echo "..." | claude -p`.

    IMPORTANTE (achado revisando o script em 20/08): a chamada aqui embaixo
    nunca teve `--model` — ou seja, toda troca de modelo feita via
    agent_versions.model no banco (inclusive tirar o CFO Synthesis de Opus)
    NUNCA teve efeito real. O agente sempre rodou no modelo padrão da sessão
    do Claude Code, não no que estava configurado. Corrigido abaixo — sem
    isso, nada do resto desta otimização de custo se sustenta."""
    user_payload = {"deal_data": deal_data}
    if other_outputs:
        user_payload["outputs_dos_outros_agentes"] = other_outputs

    full_prompt = system_prompt + "\n\n# Dados do deal\n\n" + json.dumps(user_payload, ensure_ascii=False, indent=2)

    # Retry (26/08, achado real em produção — job "opinion" quebrou com
    # o deal "Nacional", 12 meses de DRE): resposta truncada no meio,
    # nenhum dos 4 níveis de `_extract_json` conseguiu recuperar — sinal
    # de que o JSON ficou genuinamente incompleto (não é bug de parsing,
    # é a geração ter sido cortada). Isso derrubava o job inteiro depois
    # da chamada já ter sido gerada e PAGA — não repetir seria jogar
    # fora um resultado quase certo de já ter funcionado numa 2ª
    # tentativa (é comportamento raro/esporádico, não sistemático — os
    # outros 10 agentes do mesmo run passaram na 1ª tentativa).
    ultimo_erro = None
    for tentativa in range(1, 3):
        result = subprocess.run(
            ["claude", "-p", "--model", model or DEFAULT_MODEL_ALIAS, "--output-format", "json"],
            input=full_prompt,
            capture_output=True,
            text=True,
            timeout=600,  # prompts grandes com Opus podem demorar mais que os 300s originais
        )
        if result.returncode != 0:
            raise RuntimeError(
                f"claude -p saiu com código {result.returncode}\n"
                f"--- stderr ---\n{result.stderr!r}\n"
                f"--- stdout (primeiros 2000 chars) ---\n{result.stdout[:2000]!r}"
            )

        # --output-format json do Claude Code envolve a resposta num
        # envelope; o campo com o texto do agente é "result" (confirmado
        # no teste de 17/08).
        raw = json.loads(result.stdout)
        agent_text = raw.get("result", raw.get("content", result.stdout))

        parsed = _extract_json(agent_text)
        if parsed is not None:
            return parsed
        ultimo_erro = agent_text
        print(f"[aviso] tentativa {tentativa}/2 — resposta do agente não é JSON válido, tentando de novo. "
              f"Motivo de parada informado pelo Claude Code: {raw.get('stop_reason', raw.get('subtype', 'não informado'))}",
              file=sys.stderr)

    raise RuntimeError(f"Resposta do agente não é um JSON válido (2 tentativas).\nÚltima resposta bruta: {ultimo_erro[:500]}")


def _extract_json(text: str) -> dict | None:
    """Extrai JSON de uma resposta de LLM, tolerando variações comuns:
    JSON puro, cercado em ```json ... ```, cercado em ``` ... ``` simples,
    ou com texto explicativo antes/depois do bloco. Tenta na ordem do mais
    estrito pro mais tolerante.

    BUG REAL CORRIGIDO EM 26/08 (achado em produção — job 'opinion'
    falhou, jogando fora uma chamada de API inteira já paga): o passo 2
    usava uma regex não-gulosa (`\\{.*?\\}`) pra achar o bloco entre
    ```json e ``` — isso quebra com QUALQUER JSON aninhado (objeto
    dentro de objeto), porque o `.*?` para no PRIMEIRO "}" que encontra,
    que é o fechamento do objeto INTERNO, não do principal. O JSON
    capturado ficava incompleto (sem o "}" de fechamento de verdade),
    dava erro de parse, e a resposta inteira — já gerada e paga — era
    descartada. Substituído por contagem de chaves balanceada (conta
    "{"/"}" ignorando os que aparecem DENTRO de strings, pra não se
    confundir com chave sendo mencionada como texto livre num campo tipo
    "parecer_do_time") — isso encontra o "}" que fecha corretamente o
    primeiro "{", não importa quantos níveis de aninhamento existam."""
    text = text.strip()

    # 1. JSON puro
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # 2. Remove os marcadores de bloco markdown (```json / ```) do
    # início e fim, se existirem, e tenta de novo como JSON puro —
    # cobre o caso mais comum (resposta é só o bloco cercado, nada
    # antes/depois) sem depender de regex nenhuma.
    sem_cerca = re.sub(r"^```(?:json)?\s*", "", text)
    sem_cerca = re.sub(r"\s*```\s*$", "", sem_cerca).strip()
    if sem_cerca != text:
        try:
            return json.loads(sem_cerca)
        except json.JSONDecodeError:
            pass

    # 3. Contagem de chaves balanceada, a partir do primeiro "{" —
    # funciona com qualquer nível de aninhamento, e ignora "{"/"}" que
    # apareçam DENTRO de strings (texto livre), não só de objetos.
    for candidato in (sem_cerca, text):
        inicio = candidato.find("{")
        if inicio == -1:
            continue
        profundidade = 0
        dentro_de_string = False
        escapando = False
        for i in range(inicio, len(candidato)):
            ch = candidato[i]
            if escapando:
                escapando = False
                continue
            if ch == "\\":
                escapando = True
                continue
            if ch == '"':
                dentro_de_string = not dentro_de_string
                continue
            if dentro_de_string:
                continue
            if ch == "{":
                profundidade += 1
            elif ch == "}":
                profundidade -= 1
                if profundidade == 0:
                    try:
                        return json.loads(candidato[inicio:i + 1])
                    except json.JSONDecodeError:
                        break  # tenta o próximo candidato, se houver

    # 4. Último recurso: do primeiro "{" ao último "}" no texto inteiro
    # (mantido como rede de segurança final — mais frágil que o passo 3,
    # mas cobre casos exóticos que a contagem balanceada não previu).
    first, last = text.find("{"), text.rfind("}")
    if first != -1 and last != -1 and last > first:
        try:
            return json.loads(text[first:last + 1])
        except json.JSONDecodeError:
            pass

    return None


def compute_input_hash(deal_id: str, agent_version_id: str, checksum: str) -> str:
    raw = f"{deal_id}:{agent_version_id}:{checksum}"
    return hashlib.sha256(raw.encode()).hexdigest()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent", required=True)
    parser.add_argument("--deal-id", required=True)
    args = parser.parse_args()
    inicio = datetime.now(timezone.utc)  # ver comentário equivalente em run_extraction

    if args.agent == "extraction":
        run_extraction(args.deal_id)
        return

    deal_data = get_deal_data(args.deal_id)
    agent_version = get_active_agent_version(args.agent)

    input_hash = compute_input_hash(args.deal_id, agent_version["id"], deal_data["checksum"])

    # Idempotência: se já existe um run com esse hash exato, não reprocessa.
    table = "synthesis_runs" if args.agent == "cfo_synthesis" else "agent_runs"
    existing = supabase_request(
        "GET",
        f"{table}?deal_id=eq.{args.deal_id}&agent_version_id=eq.{agent_version['id']}"
        + ("" if table == "synthesis_runs" else f"&input_hash=eq.{input_hash}"),
    )
    if existing:
        print(f"[{args.agent}] já existe um run para este input — pulando (idempotência).")
        return

    # Só o conteúdo relevante vai para o prompt — id/checksum são bookkeeping
    # interno, não fazem parte do que o agente precisa analisar. A série
    # contábil crua só vai pro agente financeiro, e mesmo assim já reduzida
    # às contas relevantes (ver select_relevant_accounts) — os demais
    # recebem só o resumo estruturado.
    agent_input = {"structured": deal_data.get("structured", {})}
    if args.agent in AGENTS_NEED_RAW_SERIES:
        raw = dict(deal_data.get("raw_extracted", {}))

        # Motor determinístico (financial_engine.py + dre_balancete_parser.py):
        # o Claude não recebe mais a série crua pra "descobrir" EBITDA e
        # anomalia sozinho — código já calculou, ele só interpreta em poucas
        # palavras o que já foi achado. Corta contexto de entrada E o
        # "raciocínio" caro que o modelo fazia pra vasculhar 300+ contas.
        dre_estruturada = raw.get("dre_estruturada") or {}
        dre_hierarquia_aproximada = raw.get("dre_hierarquia_aproximada") or {}
        if dre_estruturada:
            # DRE é a fonte PRIMÁRIA — já vem categorizada pela própria
            # planilha da empresa, sem risco de duplicar hierarquia (ver
            # 21/08: o waterfall reconstruído do Balancete cru precisou de
            # 5 correções pra bater, a DRE bateu de primeira).
            primeira_dre = next(iter(dre_estruturada.values()))
            raw["ebitda_calculado"] = calcular_ebitda_de_dre(primeira_dre)
            raw["anomalias_detectadas"] = detectar_anomalias_run_rate(
                dre_linhas_para_contas(primeira_dre), top_n=10
            )
        elif dre_hierarquia_aproximada and next(iter(dre_hierarquia_aproximada.values()), {}).get("multi_entidade_ambigua"):
            # Achado real em 27/08 (deal Irko) — aba com múltiplos blocos
            # de empresa (holding) e NENHUM bloco "combinado"/"consolidado"
            # claro pra escolher sozinho (ver `_aplicar_deteccao_multi_empresa`
            # em dre_balancete_parser.py). Não calcula NADA aqui — sem
            # isso, cairia no mesmo bug que gerou o resultado errado do
            # Irko (pegar uma subsidiária isolada como se fosse o total).
            # `ebitda_calculado=None` explícito é preferível a um número
            # plausível mas errado: força o `financial_analysis`/
            # `cfo_synthesis` a tratar isso como bloqueio, não como dado.
            primeiro_arquivo = next(iter(dre_hierarquia_aproximada.values()))
            raw["ebitda_calculado"] = None
            raw["anomalias_detectadas"] = []
            raw.pop("dre_hierarquia_aproximada", None)
            raw["dre_hierarquia_nota"] = (
                "BLOQUEADO: " + primeiro_arquivo["nota"] +
                " Nenhum EBITDA/margem foi calculado a partir desta DRE — "
                "trate como dado ausente, não como zero ou não-material."
            )
        elif dre_hierarquia_aproximada:
            # Fallback (25/08) — a DRE existe e foi lida, mas a nomenclatura
            # não bateu DRE_CATEGORIAS; usa o resultado já calculado na
            # extração (`extrair_hierarquia_dre` + `calcular_resultado_de_hierarquia`,
            # ver dre_balancete_parser.py). Deliberadamente mais grosseiro
            # (não separa impostos sobre venda de impostos sobre lucro) —
            # por isso NUNCA fica em silêncio sobre isso: `fonte` e
            # `hierarquia_confiavel` vão junto pro prompt do agente, que
            # deve refletir essa incerteza na análise, não escondê-la.
            primeiro_arquivo = next(iter(dre_hierarquia_aproximada.values()))
            raw["ebitda_calculado"] = primeiro_arquivo["resultado"]
            raw["anomalias_detectadas"] = []  # não calculável sem categorias finas — não inventa
            # Remove a hierarquia BRUTA (todas as raízes, valor por mês —
            # pode ser tão grande quanto a DRE original) depois de já ter
            # extraído só o resultado agregado pequeno acima. Sem isso, a
            # hierarquia completa vazaria inteira pro prompt do agente,
            # voltando ao problema original (dump grande pra IA).
            raw.pop("dre_hierarquia_aproximada", None)
            raw["dre_hierarquia_nota"] = (
                "DRE não bateu nomenclatura conhecida — resultado aproximado por "
                "hierarquia estrutural (indentação/colunas), não pela categorização fina "
                "de sempre. Ver ebitda_calculado.fonte e .hierarquia_confiavel."
            )
        else:
            # Sem DRE própria no arquivo — cai pro Balancete cru como plano B.
            series = raw.get("series_contabeis") or []
            if series:
                raw["waterfall_calculado"] = mapear_waterfall(series)
                raw["ebitda_calculado"] = calcular_ebitda_3_camadas(
                    raw["waterfall_calculado"], lucro_liquido=None,
                    tributos_sobre_lucro=None, resultado_financeiro=None,
                )
                raw["anomalias_detectadas"] = detectar_anomalias_run_rate(series, top_n=10)

        series = raw.get("series_contabeis") or []
        if series:
            raw["contas_novas_ou_zeradas"] = detectar_contas_novas_ou_zeradas(series)
            # A série crua completa não vai mais — o Claude já recebeu o que
            # importa dela (anomalias + novas/zeradas) calculado acima.
            raw.pop("series_contabeis", None)
            raw["series_contabeis_nota"] = (
                "Série crua não incluída — EBITDA e anomalias já calculados por código "
                "(ver ebitda_calculado / anomalias_detectadas / contas_novas_ou_zeradas)."
            )

        # `dre_estruturada` (dict bruto {categoria: {mes_NN: valor}},
        # sem hierarquia nem agregação) sai do contexto — é redundante
        # com `ebitda_calculado`/`anomalias_detectadas` (já calculados
        # acima) e é o formato mais verboso dos três.
        if raw.get("dre_estruturada"):
            raw.pop("dre_estruturada", None)
            raw["dre_estruturada_nota"] = (
                "DRE bruta linha-a-linha não incluída aqui — já resumida acima em "
                "ebitda_calculado/anomalias_detectadas. Ver mini_dre/tabela_viabilidade_financeira "
                "pra números por período (agrupados por trimestre quando há muitos)."
            )

        # `mini_dre`/`tabela_viabilidade_financeira`: AGRUPADOS (não
        # removidos) quando há muitos períodos — achado real em 26/08
        # (deal "Nacional", 12 meses de DRE, Ponto 1 do Thiago: "reduzir
        # tempo e consumo"). Corrigindo uma primeira tentativa errada:
        # cheguei a remover os dois campos por completo daqui, assumindo
        # que o agente já via os mesmos números via `format_dre_table`
        # — só que esse texto (`combined_text`) só é visto pelo agente
        # `extraction`, nunca por `opinion`/`financial_analysis`/
        # `cfo_synthesis` (que são os únicos que passam por este trecho,
        # `AGENTS_NEED_RAW_SERIES`). Remover sem dar nada em troca teria
        # deixado esses 3 agentes cegos aos números detalhados da DRE —
        # pior que o problema original. Agrupar (mesma lógica já testada
        # no PPT) reduz o volume sem cegar ninguém.
        for campo in ("mini_dre", "tabela_viabilidade_financeira"):
            estrutura = raw.get(campo)
            if estrutura:
                # Reconstrói em cópias novas — nunca muta `dados` no lugar
                # (é o mesmo objeto referenciado em `deal_data`, que este
                # código só deveria LER, não alterar).
                raw[campo] = {
                    arquivo: {**dados, "linhas": agregar_linhas_por_trimestre(dados["linhas"])} if dados.get("linhas") else dados
                    for arquivo, dados in estrutura.items()
                }

        agent_input["raw_extracted"] = raw

    # CFO Synthesis precisa ver o que os outros 5 agentes concluíram pra
    # reconciliar divergências — o caminho genérico não levava isso (gap
    # encontrado revisando o script em 20/08). Busca os agent_runs do deal
    # e monta {nome_do_agente: output}.
    # NOTA: não testado contra o Supabase real ainda — a sintaxe de select
    # aninhado do PostgREST assume que agent_runs.agent_version_id ->
    # agent_versions.agent_id -> agents.id são as FKs reais; confirmar
    # antes do primeiro run em produção.
    other_outputs = None
    if args.agent == "cfo_synthesis":
        prior_runs = supabase_request(
            "GET",
            f"agent_runs?deal_id=eq.{args.deal_id}"
            "&select=output,agent_versions(agents(name))",
        )
        other_outputs = {
            r["agent_versions"]["agents"]["name"]: r["output"]
            for r in prior_runs
            if r.get("agent_versions", {}).get("agents", {}).get("name")
        }

    if args.agent == "complexity":
        # Bloco B já foi calculado na extração (run_extraction), a partir
        # do formulário processado 100% em código — aqui só copiamos pra
        # agent_runs, sem recalcular e sem chamada de rede.
        output = deal_data.get("raw_extracted", {}).get("complexidade_operacional_calculada")
        if output is None:
            # Sem formulário reconhecido nesse deal — fallback pro motor
            # antigo (mais simples), lendo os campos que a extração possa
            # ter preenchido via LLM em vez do formulário.
            resultado = classificar_complexidade(deal_data.get("structured", {}))
            output = resultado.to_agent_run_output()
    elif args.agent == "viabilidade_financeira":
        # Idem — Bloco A já calculado na extração, zero chamada de rede.
        output = deal_data.get("raw_extracted", {}).get("viabilidade_financeira_calculada")
        if output is None:
            output = {"classificacao": "Dados insuficientes", "motivo": "Formulário não reconhecido nesse deal — sem Bloco A calculado."}
    elif args.agent == "integration_risks":
        # Aposentado em 24/08, mesmo padrão do operational_risks: os 2
        # booleanos que este agente usava já existem no Bloco B, de graça
        # — confirmado lendo o prompt real dele no Supabase antes de
        # decidir. Ver avaliar_riscos_integracao() em regras_negocio.py
        # (inclui nota sobre o checklist de due diligence não migrado).
        output = deal_data.get("raw_extracted", {}).get("riscos_integracao_calculados")
        if output is None:
            output = {"nivel_risco": "não avaliado", "riscos_integracao": ["Formulário não reconhecido nesse deal."]}
    elif args.agent == "operational_risks":
        # Aposentado em 24/08: todo critério numérico que este agente
        # calculava (concentração top 10, headcount vs. carteira, sistema
        # do cliente, inadimplência) já existe no Bloco A/B, de graça —
        # confirmado lendo o prompt real dele no Supabase antes de decidir.
        # Ver avaliar_riscos_operacionais() em regras_negocio.py.
        output = deal_data.get("raw_extracted", {}).get("riscos_operacionais_calculados")
        if output is None:
            output = {"nivel_concentracao": "não avaliado", "riscos_operacionais": ["Formulário não reconhecido nesse deal."]}
    else:
        output = call_claude(
            agent_version["system_prompt"], agent_input, other_outputs,
            model=agent_version.get("model"),
        )

    row = {
        "deal_id": args.deal_id,
        "agent_version_id": agent_version["id"],
        "status": "completed",
        "output": output,
        "confidence": output.get("confidence"),
        "started_at": inicio.isoformat(),
    }
    if table == "agent_runs":
        row["deal_data_id"] = deal_data["id"]
        row["input_hash"] = input_hash
    else:
        row["recommendation"] = output.get("recommendation")

    supabase_request("POST", table, row)
    print(f"[{args.agent}] concluído e gravado em {table}.")

    # O CFO Synthesis é o último passo da análise em si (Output Generator é
    # depois, separado) — é o momento certo de marcar o deal como concluído.
    if args.agent == "cfo_synthesis":
        supabase_request("PATCH", f"deals?id=eq.{args.deal_id}", {"status": "completed"})
        print("deal marcado como completed.")


if __name__ == "__main__":
    main()
