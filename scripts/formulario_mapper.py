"""
formulario_mapper.py — traduz as respostas cruas do formulário (export tipo
Microsoft/Google Forms) pros parâmetros que `regras_negocio.py` espera.

Casamento de coluna é por PALAVRA-CHAVE (regex), não texto exato — resiliente
a pequenas mudanças de enunciado (plural, acento, ordem de palavra) sem
custar nada (é regex, não IA). Confirmado com o Thiago em 21/08: o
cabeçalho é fixo hoje, mas pode ganhar campos novos no futuro — por isso
essa resiliência, de graça, em vez de casar por índice de coluna ou texto
exato.
"""
from __future__ import annotations

import re

MUNICIPIOS_GRANDE_SP = {
    "são paulo", "sao paulo", "guarulhos", "osasco", "santo andré", "santo andre",
    "são bernardo do campo", "sao bernardo do campo", "são caetano do sul", "sao caetano do sul",
    "diadema", "barueri", "carapicuíba", "carapicuiba", "cotia", "embu das artes",
    "itapevi", "jandira", "taboão da serra", "taboao da serra", "mauá", "maua",
    "ribeirão pires", "ribeirao pires", "suzano", "mogi das cruzes", "itaquaquecetuba",
}

SEGMENTOS_SENSIVEIS_REGEX = re.compile(r"(?i)ind[úu]stria|associa[çc][ãa]o|hospital|institui[çc][ãa]o de ensino|ensino")

# Nome canônico -> padrões que podem aparecer no cabeçalho da pergunta.
# Cada campo tenta os padrões em ordem; o primeiro que bater na coluna vence.
CAMPOS_FORMULARIO = {
    "regime_tributario": [r"regime\s*tribut[áa]rio"],
    "faturamento_mensal": [r"faturamentos?\s*mensa(l|is)"],
    "folha_informada": [r"custos?\s*de\s*folha\s*operacional", r"custo.*folha"],
    "custo_sistemas": [r"custos?\s*mensa(l|is)\s*em\s*sistemas", r"custo.*sistemas"],
    "numero_clientes": [r"n[úu]mero.*clientes.*carteira"],
    "numero_colaboradores": [r"pessoas\s*envolvidas\s*nas\s*opera[çc][õo]es"],
    "concentracao_top10_pct": [r"faturamento\s*top\s*10", r"top\s*10\s*maiores\s*clientes"],
    "composicao_carteira": [r"composi[çc][ãa]o\s*da\s*carteira"],
    "perfil_restrito": [r"criptomoeda", r"perfis\s*restritos", r"se\s*enquadram\s*nos\s*seguintes\s*perfis"],
    "endereco": [r"endere[çc]o\s*do\s*escrit[óo]rio"],
    "sistema_erp": [r"sistema\s*de\s*erp\s*cont[áa]bil", r"erp\s*cont[áa]bil"],
    # Achado real em 26/08: quando "sistema_erp" vem genérico ("Outro",
    # sem dizer qual), este campo tem o nome de verdade — sem ele, o red
    # flag de ERP pouco difundido citava "Outro" como se fosse o nome do
    # sistema, o que não ajuda ninguém a saber qual sistema é.
    "sistemas_cliente_texto": [r"quais\s*sistemas\s*do\s*cliente\s*s[ãa]o\s*utilizados"],
    "sistema_financeiro": [r"sistema\s*financeiro"],
    "opera_sistema_cliente_exceto_omie": [
        r"executa\s*atividades\s*operacionais.*sistema\s*do\s*cliente",
        r"atividades\s*operacionais\s*diretamente\s*no\s*sistema\s*do\s*cliente",
    ],
    "infraestrutura": [r"infraestrutura\s*do\s*sistema"],
    "outsourcing_sistemas_pct": [r"faturamento\s*mensal\s*sistemas\s*do\s*cliente"],
    "outsourcing_pessoas_pct": [r"pessoas\s*alocadas\s*no\s*cliente"],
    "outsourcing_sistemas_clientes_qtd": [r"clientes\s*atendidos\s*no\s*caso\s*acima"],
    "outsourcing_pessoas_clientes_qtd": [r"clientes\s*que\s*possuem\s*pessoas\s*alocadas"],
    "inadimplencia_pct": [r"inadimpl[êe]ncia"],
    "churn_pct": [r"churn"],
}


def _pct_para_100(valor) -> float | None:
    """Normaliza porcentagem pra escala 0-100 — o motor de regras espera
    0-100 (ex.: 8 = 8%, não 0.08).

    BUG REAL CORRIGIDO EM 25-26/08: esta função multiplicava por 100
    incondicionalmente, assumindo que o formulário sempre grava
    porcentagem como fração (0.2 = 20%). Testando contra os 2
    formulários reais já processados (BPO Innova e CSF Hotelaria — 10
    campos percentuais no total), NENHUM vem como fração — todos já
    vêm como número inteiro direto (90 = 90%, 8 = 8%, até 1 = 1%).
    Multiplicar sempre por 100 dava valores 100x maiores em TODOS os
    casos reais (churn virando "800%", concentração virando "9000%" —
    o próprio agente `cfo_synthesis` sinalizou isso como
    "matematicamente impossível" num teste real, 26/08). Isso não é um
    bug introduzido agora — afeta qualquer deal já processado antes
    desta correção.

    Uma primeira tentativa de correção multiplicava só quando o valor
    vinha <=1 (assumindo fração só nesse caso) — mas isso quebrou um
    caso real também: inadimplência=1.0 (significa 1%) virava 100.0.
    Como os 10 valores reais já vistos NUNCA usam fração — nem os
    pequenos (inadimplência=1, outsourcing=2) — a correção mais segura,
    apoiada na evidência real disponível, é não multiplicar nunca. Se
    algum dia aparecer um formulário real que genuinamente grave como
    fração, essa é a função a revisar de novo — com o exemplo real em
    mãos, não por suposição."""
    if valor is None:
        return None
    return round(float(valor), 4)


def _detectar_fora_grande_sp(endereco: str | None) -> bool | None:
    if not endereco:
        return None
    texto = endereco.strip().lower()
    return not any(municipio in texto for municipio in MUNICIPIOS_GRANDE_SP)


def _normalizar_hospedagem(infra: str | None) -> str | None:
    if not infra:
        return None
    texto = infra.strip().lower()
    if "fisico" in texto or "físico" in texto or "local" in texto or "on-premise" in texto or "on premise" in texto:
        return "servidor_fisico"
    if "web" in texto or "nuvem" in texto or "cloud" in texto or "saas" in texto:
        return "nuvem_ou_web"
    return None


def localizar_colunas(cabecalhos: list) -> dict:
    """Constrói {campo_canonico: indice_da_coluna} varrendo o cabeçalho real
    da planilha com os padrões de CAMPOS_FORMULARIO. Roda uma vez por
    arquivo (custo desprezível) — depois disso, extrair cada linha de
    resposta é indexação direta, sem regex de novo."""
    indice = {}
    cabecalhos_norm = [(_norm(h) if h else "") for h in cabecalhos]
    for campo, padroes in CAMPOS_FORMULARIO.items():
        for padrao in padroes:
            regex = re.compile(padrao, re.IGNORECASE)
            achou = next((i for i, h in enumerate(cabecalhos_norm) if regex.search(h)), None)
            if achou is not None:
                indice[campo] = achou
                break
    return indice


def _norm(texto: str) -> str:
    return str(texto).replace("\xa0", " ").strip()


def extrair_respostas_linha(row: tuple, indice_colunas: dict) -> dict:
    """Extrai uma linha de resposta usando o índice já construído por
    `localizar_colunas` — {campo_canonico: valor_bruto}. Campo cujo
    cabeçalho não foi encontrado no arquivo vem como None (nunca quebra)."""
    return {campo: (row[col] if col < len(row) else None) for campo, col in indice_colunas.items()}


LIMIAR_CONFIANCA_FORMULARIO = 8  # de ~18 campos conhecidos — abaixo disso, não confia que é o formulário


def detectar_e_extrair_formulario(file_bytes: bytes) -> dict | None:
    """Tenta reconhecer o arquivo como o formulário do parceiro (cabeçalho
    fixo, uma linha de resposta) e, se reconhecer com confiança, já extrai
    a linha 2 (primeira resposta) em campos canônicos. Retorna None se não
    bater confiança suficiente — nesse caso o chamador trata como arquivo
    genérico (Balancete/DRE/outro), sem travar em nada."""
    try:
        from openpyxl import load_workbook
        import io as _io
        wb = load_workbook(_io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return None

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        linhas = ws.iter_rows(min_row=1, max_row=2, values_only=True)
        try:
            cabecalhos = next(linhas)
        except StopIteration:
            continue
        indice = localizar_colunas(list(cabecalhos))
        if len(indice) < LIMIAR_CONFIANCA_FORMULARIO:
            continue
        try:
            primeira_resposta = next(linhas)
        except StopIteration:
            continue
        return extrair_respostas_linha(primeira_resposta, indice)
    return None


def _pct_clientes(qtd_clientes, total_clientes) -> float | None:
    """Outsourcing como % de CLIENTES envolvidos sobre a carteira total —
    não % de faturamento. Achado real em 26/08 (Thiago, revisando a
    matriz de regras): o critério 8/9 do Bloco B ("Outsourcing — Pessoas
    Alocadas"/"Sistemas do Cliente") deveria usar a proporção de
    CLIENTES com outsourcing, não a fração da RECEITA que eles
    representam — são perguntas diferentes, e usar a errada muda o
    resultado de forma grave (no BPO Innova real: outsourcing de
    pessoas por % de faturamento = 60% (>15%, "Alta"); por % de
    clientes = 3/31 ≈ 9,7% ("Baixa") — classificação oposta)."""
    if qtd_clientes is None or not total_clientes:
        return None
    return round(float(qtd_clientes) / float(total_clientes) * 100, 2)


def _valor_erp_final(erp_bruto: str | None, sistemas_cliente: str | None) -> str | None:
    """Resolve o nome de verdade do sistema — quando "Sistema de ERP
    contábil" veio genérico ("Outro", sem dizer qual), usa o campo
    "Quais sistemas do cliente são utilizados..." como fallback. Achado
    real em 26/08: sem isso, o red flag de ERP pouco difundido citava
    "Outro" como se fosse o nome do sistema, o que não identifica nada
    pro time de due diligence."""
    if erp_bruto and erp_bruto.strip().lower() not in ("outro", "outros", ""):
        return erp_bruto
    return sistemas_cliente or erp_bruto


def mapear_formulario(respostas_canonicas: dict, rbt12_real: float | None = None) -> dict:
    """`respostas_canonicas` é o dict {campo_canonico: valor} que
    `extrair_respostas_linha` já produziu. Retorna pronto pra **desempacotar**
    direto nos kwargs de `avaliar_viabilidade_financeira` e
    `avaliar_complexidade_operacional`.

    RBT12 (Receita Bruta em 12 Meses, usada no Simples Nacional): preferimos
    `rbt12_real` — soma de 12 meses de receita bruta já extraída da DRE, bem
    mais confiável que aproximar a partir de 1 mês do formulário (decisão do
    Thiago em 21/08: "pegaremos na DRE"). Só cai pra `faturamento_mensal × 12`
    se não houver DRE nesse deal — e nesse caso o aviso deixa isso explícito."""
    g = respostas_canonicas.get

    faturamento_mensal = g("faturamento_mensal")
    if rbt12_real is not None:
        rbt12 = rbt12_real
        avisos_mapeamento = []
    else:
        rbt12 = faturamento_mensal * 12 if faturamento_mensal is not None else None
        avisos_mapeamento = [
            "RBT12 aproximado como faturamento_mensal × 12 (nenhuma DRE disponível "
            "nesse deal pra usar a receita bruta real dos 12 meses) — confirmar se "
            "o mês informado é representativo."
        ]
    composicao_carteira = g("composicao_carteira") or ""
    segmentos_sensiveis_presentes = bool(SEGMENTOS_SENSIVEIS_REGEX.search(composicao_carteira))
    perfil_restrito_raw = (g("perfil_restrito") or "").strip().lower()

    # Sistema financeiro: hoje não existe pergunta dedicada, mas a pergunta
    # "opera no sistema do cliente, exceto Omie/Conta Azul?" já resolve isso
    # na prática (confirmado com o Thiago em 21/08): "Sim" = usa sistema do
    # cliente que NÃO é Omie/Conta Azul -> não é Omie. "Não" = só opera com
    # Omie/Conta Azul (ou nenhum) -> é Omie. Se um dia existir uma pergunta
    # dedicada de "sistema financeiro", ela tem prioridade (mais direta).
    sistema_financeiro_texto = g("sistema_financeiro")
    if sistema_financeiro_texto:
        sistema_financeiro_e_omie = "omie" in sistema_financeiro_texto.strip().lower()
    else:
        opera_cliente_raw = (g("opera_sistema_cliente_exceto_omie") or "").strip().lower()
        if opera_cliente_raw == "sim":
            sistema_financeiro_e_omie = False
        elif opera_cliente_raw in ("não", "nao"):
            sistema_financeiro_e_omie = True
        else:
            sistema_financeiro_e_omie = None

    return {
        # --- Bloco A ---
        "regime_tributario": g("regime_tributario"),
        "faturamento_mensal": faturamento_mensal,
        "folha_informada": g("folha_informada"),
        "custo_sistemas": g("custo_sistemas"),
        "rbt12": rbt12,
        "inadimplencia_media_pct": _pct_para_100(g("inadimplencia_pct")),
        "churn_medio_pct": _pct_para_100(g("churn_pct")),
        "perfil_restrito_presente": (perfil_restrito_raw == "sim") if perfil_restrito_raw else None,

        # --- Bloco B ---
        "localizacao_fora_grande_sp": _detectar_fora_grande_sp(g("endereco")),
        "numero_clientes": g("numero_clientes"),
        "numero_colaboradores": g("numero_colaboradores"),
        "concentracao_top10_pct": _pct_para_100(g("concentracao_top10_pct")),
        "segmentos_sensiveis_presentes": segmentos_sensiveis_presentes,
        "sistema_financeiro_e_omie": sistema_financeiro_e_omie,
        "sistema_utilizado_texto": _valor_erp_final(g("sistema_erp"), g("sistemas_cliente_texto")),
        "sistema_hospedagem": _normalizar_hospedagem(g("infraestrutura")),
        "outsourcing_pessoas_pct": _pct_clientes(g("outsourcing_pessoas_clientes_qtd"), g("numero_clientes")),
        "outsourcing_sistemas_pct": _pct_clientes(g("outsourcing_sistemas_clientes_qtd"), g("numero_clientes")),
        # Dado auxiliar (não usado no motor de regras) — % de FATURAMENTO
        # que esses clientes com outsourcing representam, pra contexto
        # qualitativo. Guardado, não descartado, mesmo não sendo mais a
        # fonte do critério Bloco B.
        "outsourcing_pessoas_pct_faturamento": _pct_para_100(g("outsourcing_pessoas_pct")),
        "outsourcing_sistemas_pct_faturamento": _pct_para_100(g("outsourcing_sistemas_pct")),

        "avisos_mapeamento": avisos_mapeamento,
    }
